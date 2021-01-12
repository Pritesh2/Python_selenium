# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and setti

import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import openpyxl
import Utils

def write_xl():
    path="/home/pritesh/Desktop/SEBI/Veritas2.xlsx"
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name('Sheet1')

    for r in range(1,10):
        for c in range(1,8):
            sheet.cell(row=r,column=c).value="Hello"

    workbook.save(path)

def read_xl():
    path="/home/pritesh/Desktop/SEBI/Veritas2.xlsx"

    workbook=openpyxl.load_workbook(path)

    sheet=workbook.get_sheet_by_name('Sheet1')

    ## find no of rows and column

    r=sheet.max_row
    c=sheet.max_column
    print(r)
    print(c)

    for i in range(1,r+1):
        #if(str(sheet.cell(row=i,column=1).value)=='Email-ID'):
            #print(dtype(sheet.cell(row=1,column=)))
            for j in range(1,c+1):
                if(sheet.cell(row=1,column=j).value=='Email-ID'):
                    print(sheet.cell(row=i,column=j).value,end=" ")
            print()

#path="/home/pritesh/Desktop/SEBI/veritasexl.xlsx"

def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi from Pritesh, {name}')  # Press Ctrl+F8 to toggle the breakpoint.

    #driver=webdriver.Firefox(executable_path='/home/pritesh/Desktop/Veritas_Python/geckodriver-v0.27.0-linux64/geckodriver')

    driver=webdriver.Chrome(executable_path="/home/pritesh/Desktop/Veritas_Python/chromedriver_linux64/chromedriver")
    driver.get('https://www.spoj.com/users/probito/')
    links=driver.find_elements(By.TAG_NAME,"a")
    for link in links:
        print(link.text)

    #print(len(links))

# Press the green button in the gutter to run the script.

def f1():
    path="/home/pritesh/Desktop/SEBI/Veritas2.xlsx"


    driver = webdriver.Firefox(executable_path='/home/pritesh/Desktop/Veritas_Python/geckodriver-v0.27.0-linux64/geckodriver')

    driver.get("https://github.com/login")
    driver.maximize_window()

    r=Utils.getRowCount(path,"Sheet2")
    c=Utils.getColumnCount(path,"Sheet2")

    print(r)
    print(c)

    for i in range(1,r+1):
        username=Utils.readData(path,"Sheet2",i,1)
        password=Utils.readData(path,"Sheet2",i,2)
        #print(password)
        driver.find_element_by_name("login").send_keys(username)
        driver.find_element_by_name("password").send_keys(password)

        driver.find_element_by_name("commit").click()

        if driver.title=="GitHub":
            #print("Passed")
            Utils.writeData(path,"Sheet2",i,3,"Passed")

        else:
            #print("Invalid")
            Utils.writeData(path,"Sheet2",i,3,"Failed")

        driver.get("https://github.com/login")
    #workbook.save(file)
    driver.close()


if __name__ == '__main__':
    f1()
    ##print_hi('PyCharm')
    ##read_xl()
    #write_xl()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
