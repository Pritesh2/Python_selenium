import openpyxl

def getColumnCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return sheet.max_column

def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return sheet.max_row

def readData(file,sheetName,rowno,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rowno,column=columnno).value

def writeData(file,sheetName,rowno,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowno,column=columnno).value=data
    workbook.save(file)
    print(sheet.cell(row=rowno,column=columnno).value)

   # workbook = openpyxl.load_workbook(path)
    #sheet = workbook.get_sheet_by_name('Sheet1')


    #        sheet.cell(row=r, column=c).value = r * c

    #workbook.save(path)